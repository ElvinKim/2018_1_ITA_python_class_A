from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()

ws = wb.active
w1 = wb.create_sheet("sheet1")
w2 = wb.create_sheet("sheet2")
ws.title = "new sheet"

ws["A1"] = 10
ws.cell(row=2, column=1, value=20)

# 실습
for row in range(1, 11):
    for col in range(1, 11):
        ws.cell(row=row, column=col, value=(row-1) * 10 + col)

# 여러 셀에 접근하기
for col_c in ws["C"]:
    print(col_c)

print("-" * 20)

for data in ws["A:C"]:
    for d in data:
        print(d.value)
    print("=" * 20)

print("-" * 20)

for row_2 in ws[2]:
    print(row_2.value)

for data in ws[1:3]:
    for d in data:
        print(d.value)
    print("=" * 20)

print("-" * 20)
for row in ws.iter_rows(min_row=2, max_row=8, min_col=2, max_col=5):
    print(row)

print("-" * 20)
for col in ws.iter_cols(min_row=2, max_row=8, min_col=2, max_col=5):
    print(col)

# Save & Load
# wb.save("test.xlsx")

loaded_wb = load_workbook("test.xlsx")
loaded_ws = loaded_wb.get_sheet_by_name("new sheet")
print(loaded_ws["A1"].value)
