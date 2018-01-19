import openpyxl

word_wb = openpyxl.load_workbook("words.xlsx")
word_ws = word_wb.get_sheet_by_name("words")

# Problem 1
problem1_ws = word_wb.create_sheet("problem 1")

for row_i, word in enumerate(word_ws["A"]):
    problem1_ws.cell(row=row_i + 1, column=1, value=word.value)

# Problem 2
problem2_ws = word_wb.create_sheet("problem 2")

for row_i, row in enumerate(word_ws.iter_rows()):
    problem2_ws.cell(row=row_i + 1, column=1, value=row[1].value)
    problem_en_word = row[0].value[0] + "_" * (len(row[0].value) - 1)
    problem2_ws.cell(row=row_i + 1, column=2, value=problem_en_word)

# Problem 3
problem3_ws = word_wb.create_sheet("problem 3")

for row_i, row in enumerate(word_ws.iter_rows()):
    problem3_ws.cell(row=row_i + 1, column=1, value=row[4].value)

    word_lst = []

    for word in row[3].value.split():
        if word[-1] == "," or word[-1] == ".":
            word = word[0] + "_" * (len(word) - 2) + word[-1]
        else:
            word = word[0] + "_" * (len(word) - 1)
        word_lst.append(word)

    problem3_ws.cell(row=row_i + 1, column=2, value=" ".join(word_lst))
word_wb.save("words.xlsx")



